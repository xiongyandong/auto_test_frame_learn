from openpyxl import load_workbook
from Base.setting import Setting
import os


class LoadData:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open_excel(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        return sheet

    def get_header(self):
        header = []
        sheet = self.open_excel()
        max_column = sheet.max_column
        for j in range(1, max_column + 1):
            header.append(sheet.cell(1, j).value)
        return header

    def load_excel(self, mode='all'):
        sheet = self.open_excel()
        header = self.get_header()
        max_row = sheet.max_row
        max_column = sheet.max_column
        tests_data = []
        for row in range(2, max_row + 1):
            temporary_list = []
            for column in range(1, max_column + 1):
                cell_value = sheet.cell(row, column).value
                temporary_list.append(cell_value)
            test_data = {header[0]: temporary_list[0], header[1]: temporary_list[1], header[2]: temporary_list[2],
                         header[3]: temporary_list[3], header[4]: eval(temporary_list[4]), header[5]: temporary_list[5],
                         header[6]: temporary_list[6]}
            tests_data.append(test_data)
        if mode == 'all':
            final_data = tests_data
        else:  # [1,2,3,4]
            final_data = []
            for item in tests_data:
                if item['case_id'] in eval(mode):
                    final_data.append(item)

        return final_data
