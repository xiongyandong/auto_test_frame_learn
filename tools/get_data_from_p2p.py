from openpyxl import load_workbook
from Base.setting import Setting
import os
import re
import json
from tools.log import MyLog
case_variable_path = os.path.join(Setting.BaseDir, r'data\p2pData\case_variable.json')


class LoadData:
    def __init__(self, file_name, module=None):
        self.file_name = file_name
        self.module = module
        # self.initial_phone = phone_num
        self.sheet_name = list(module.keys())
        self.case_variable = ReadGlobalVariable(case_variable_path).get_global_variable()

    def open_excel(self):
        work_sheet = []
        wb_all = load_workbook(self.file_name)
        for sheet in self.sheet_name:
            wb = wb_all[sheet]
            mode = self.module[sheet]
            work_sheet.append([wb, mode])
        return work_sheet

    def get_header(self):
        header = []
        work_sheet = self.open_excel()
        max_column = work_sheet[0][0].max_column
        for j in range(1, max_column + 1):
            header.append(work_sheet[0][0].cell(1, j).value)
        return header

    def load_excel(self):
        work_sheet = self.open_excel()
        header = self.get_header()
        tests_data = []
        for sheet in work_sheet:
            max_row = sheet[0].max_row
            max_column = sheet[0].max_column
            mode = sheet[1]
            single_data = []
            for row in range(2, max_row + 1):
                temporary_list = []
                for column in range(1, max_column):
                    cell_value = sheet[0].cell(row, column).value
                    temporary_list.append(cell_value)
                test_data = {header[0]: temporary_list[0], header[1]: temporary_list[1], header[2]: temporary_list[2],
                             header[3]: temporary_list[3], header[4]: eval(temporary_list[4]),
                             header[5]: temporary_list[5],
                             header[6]: temporary_list[6]}
                single_data.append(test_data)
            if mode == 'all':
                tests_data.extend(single_data)
            else:  # [1,2,3,4]
                for item in single_data:
                    if item['case_id'] in mode:
                        tests_data.append(item)
        for case_data in tests_data:
            self.to_replace_variable(case_data, self.case_variable)
        return tests_data

    def to_replace_variable(self, case_data, global_variable_data):
        for key, value in case_data.items():
            if type(value) == dict:
                self.to_replace_variable(value, global_variable_data)
            elif type(value) == list:
                for i in value:
                    if type(i) == dict:
                        self.to_replace_variable(i, global_variable_data)
            else:
                if '$' in str(value):
                    key_variable = value[1:]
                    try:
                        case_data[key] = global_variable_data[key_variable]
                    except KeyError as e:
                        MyLog().my_log('找不到变量{}'.format(e), 'ERROR')
                        raise e

    def write_back(self, sheet_name, i, j, value):
        wb = load_workbook(self.file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, j).value = value
        wb.save(self.file_name)


class ReadGlobalVariable:
    def __init__(self, path):
        self.file = open(path, 'r')

    def get_global_variable(self):
        config_table_data = json.load(self.file)
        return config_table_data


if __name__ == '__main__':
    dic_str = '{"num": "136154", "data": [{"id": "$id", "tel": "$tel"}, {"id": "$id2", "tel": "$admin_tel"}]}'
    variable_list = re.findall('"\$(.*?)"', dic_str)
    print(variable_list)






